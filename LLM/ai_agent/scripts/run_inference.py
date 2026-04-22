"""
Script inference chính:
  Đọc question.xlsx → chạy pipeline từng row → ghi output/result.json

Chạy: python scripts/run_inference.py
       python scripts/run_inference.py --input path/to/question.xlsx
"""
import sys
import time
import argparse
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

import openpyxl

from src.intent.classifier import IntentClassifier
from src.retrieval.vector_store import VectorStore
from src.retrieval.bm25_store import BM25Store
from src.retrieval.reranker import Reranker
from src.llm.model import QwenModel
from src.pipeline.rag_pipeline import RAGPipeline
from src.pipeline.direct_pipeline import DirectPipeline
from src.utils.formatter import format_output, write_results

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).parent.parent
INDEX_DIR     = BASE_DIR / "data" / "index"
FAISS_DIR     = INDEX_DIR / "faiss"
BM25_PATH     = INDEX_DIR / "bm25.pkl"
DEFAULT_INPUT = BASE_DIR.parent / "public_test_data" / "public_test_data" / "question.xlsx"
OUTPUT_PATH   = BASE_DIR / "output" / "result.json"


def load_questions(xlsx_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        record = dict(zip(headers, row))
        record['_row_id'] = str(i)
        rows.append(record)
    return rows


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--input', default=str(DEFAULT_INPUT))
    parser.add_argument('--output', default=str(OUTPUT_PATH))
    args = parser.parse_args()

    # ── Kiểm tra index ──────────────────────────────────────────────────────────
    if not FAISS_DIR.exists() or not BM25_PATH.exists():
        print("Index chưa tồn tại. Chạy scripts/build_chunks.py rồi scripts/build_index.py trước!")
        sys.exit(1)

    # ── Load các component (chỉ load 1 lần) ────────────────────────────────────
    print("=== Loading components ===")
    llm           = QwenModel()
    vector_store  = VectorStore(faiss_dir=str(FAISS_DIR))
    bm25_store    = BM25Store(bm25_pkl_path=str(BM25_PATH))
    reranker      = Reranker()
    classifier    = IntentClassifier()

    rag_pipeline    = RAGPipeline(vector_store, bm25_store, reranker, llm)
    direct_pipeline = DirectPipeline(llm)

    # ── Load questions ──────────────────────────────────────────────────────────
    questions = load_questions(args.input)
    print(f"\n=== Bắt đầu inference: {len(questions)} câu hỏi ===\n")

    results = []
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    for i, row in enumerate(questions, 1):
        question = str(row.get('Question', ''))
        options  = {
            'A': str(row.get('A', '')),
            'B': str(row.get('B', '')),
            'C': str(row.get('C', '')),
            'D': str(row.get('D', '')),
        }
        note     = row.get('note') or row.get('Note')
        row_id   = row.get('id') or row['_row_id']

        t0 = time.time()

        # Intent
        intent = classifier.predict(question)

        # Pipeline
        if intent == 'call_document':
            note_str = str(note) if note else None
            raw_answer = rag_pipeline.run(question, options, note=note_str)
        else:
            raw_answer = direct_pipeline.run(question)

        elapsed = time.time() - t0

        result = format_output(
            row_id=str(row_id),
            function_code=intent,
            raw_answer=raw_answer,
            time_response=elapsed,
        )
        results.append(result)

        print(f"[{i}/{len(questions)}] id={row_id} | {intent} | ans={result['function_result']} | {elapsed:.2f}s")

    write_results(results, args.output)
    print(f"\nHoàn thành. Output: {args.output}")


if __name__ == "__main__":
    main()
